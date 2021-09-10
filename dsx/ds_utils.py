import os
import sys
import matplotlib as mpl
import pandas as pd
import warnings


import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns


from joblib import dump
from collections import Iterable
from typing import Union
import inspect



try:
	ip = get_ipython()
	if ip.has_trait('kernel'):
		# To allow multiple outputs from one cell
		from IPython.core.interactiveshell import InteractiveShell
		InteractiveShell.ast_node_interactivity = "all"

		# Pandas Table Config
		pd.set_option('display.html.table_schema', True)
		pd.set_option('display.max_rows', 250)
		print("Package loaded in Notebook Mode")
	else:
		# Else Not Imported in Jupyter Notebook or Lab
		mpl.use("module://backend_interagg")
		print("Package loaded in Non-Notebook Mode | mpl.use('module://backend_interagg')")

except:
	print('Kernel not detected. '
	      'Please use dsx.set_ipython() static method in the class to use interactive mode.'
	      'Interactive Mode is not recommended for scripts')
	pass



# Matplotlib Config
plt.style.use('fivethirtyeight')
plt.rc('figure', figsize=(16,9))
#sns.set_style('darkgrid')
#sns.set_context(context={'figure.figsize': (16,6)})


@pd.api.extensions.register_dataframe_accessor("ds")
class dsx(object):
	"""
		The **dsx** module (same name but not to confuse with the package name) contains a collection of wrapper functions to simplify common operations in data analytics tasks.
		The core module ds_utils (data science utilities) is designed to work with DataFrame in Pandas to simplify common tasks
	"""
	backup_repo = {}


	# Constructor
	def __init__(self, pandas_obj):
		self._obj = pandas_obj


	# region << Class Variable >>
	'''
	These are needed.
	Although class variables can be directly use in the @classmethod without defined here,
	but if the variable needs to be set to default value, the variable should be 'declared' here.
	'''
	dir_project = None
	dir_data = None
	dir_temp = None

	path_chrome = None
	if os.name == 'nt':
		path_chrome = 'C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s'
	elif os.name == 'posix':
		path_chrome = '/mnt/c/Program Files (x86)/Google/Chrome/Application/chrome.exe'

	path_graphviz = None
	if os.name == 'nt':
		path_graphviz = 'C:/Program Files (x86)/Graphviz2.38/bin/'
	elif os.name == 'posix':
		path_graphviz = '/mnt/c/Program Files (x86)/Graphviz2.38/bin/'


	@property
	def len(self):
		return len(self._obj)


	# endregion << Property >>


	def stdcols(self, inplace=True, camel=False):
		"""
			To standardize the names of all columns, to be compatible with iPython.
			This method removes space and special characterss in the column names.
			After standardized, the column names can be used as attribute of the DataFrame (with autocomplete) in iPython

		Parameters
		----------
		inplace: bool

		camel: bool


		Returns
		-------
		pandas.core.frame.DataFrame
			Only when inplace parameter is set to False
		"""
		df_input = self._obj
		if inplace:
			df = df_input
		else:
			df = df_input.copy()
		import re
		pat = re.compile(r'[\W]+')
		df.columns = df.columns.str.replace(pat, '_')
		df.columns = df.columns.str.strip('_')
		pat_Multi_Underscore = re.compile(r'[_]{2,}')
		df.columns = df.columns.str.replace(pat_Multi_Underscore, '_')
		if camel == True:
			df.columns = df.columns.str.title()

		if inplace:
			df_input = df.copy()
		else:
			return df


	def info(self):
		"""
		To generate the meta-data of the DataFrame.
		Meta-data includes the following:
		- Column Names
		- Missing Count
		- Missing Percentage
		- Unique Value Count (nunique)
		- Unique Value Percentage

		Returns
		-------
		pandas.core.frame.DataFrame
		"""

		cols_Df = pd.DataFrame(self._obj.columns.tolist())
		cols_Df = cols_Df.reset_index()
		cols_Df.columns = ['ColIndex', 'Col_Name']
		report_missing = self.isnull_list()
		cols_Df = cols_Df.merge(report_missing, 'left', 'Col_Name')

		report_nunique = self.nunique()
		cols_Df = cols_Df.merge(report_nunique, 'left', 'Col_Name')

		report_type = pd.DataFrame(self._obj.dtypes).reset_index()
		report_type.columns = ['Col_Name', 'Data_Type']
		report_type.Data_Type = report_type.Data_Type.map(lambda x: str(x))
		cols_Df = cols_Df.merge(report_type, 'left', 'Col_Name')
		return cols_Df


	def duplicated(self, colname_list: Union[str, list], return_dups=False, keep:bool=False) -> int:
		"""
		To count the duplicated rows, given a list of columns that contain the unique key.

		Parameters
		----------
		colname_list: Union[str, list]

		return_dups: bool, optional
			Default = False
			Set to True to return a tuple containing (count, df_duplicates).

		keep: bool, optional

		Returns
		-------
		Number of Duplicated Rows: int
		"""
		if isinstance(colname_list, list) == False:
			colname_list = [colname_list]

		df = self._obj
		if return_dups:
			dff = df[df.duplicated(subset=colname_list, keep=keep)]
			return (len(dff), dff)
		else:
			return len(df[df.duplicated(subset=colname_list, keep=keep)])


	def isnull(self, colname: str) -> tuple:
		"""
		Count the rows (and the %) of missing values in the specified column

		Parameters
		----------
		colname: str
			Single column name

		Returns
		-------
		(Count of Missing Rows, Percentage of Missing Rows): tuple
		"""

		df = self._obj
		temp = df[df[colname].isnull()]
		return (len(temp), len(temp) / len(df))


	def isnull_list(self, col_names_list=None) -> pd.core.frame.DataFrame:
		"""
		Generate a report of cases with missing values

		Parameters
		----------
		col_names_list: list, optional
			List of columns to be included in the report.
			If not specified, all columns will be used.


		Returns
		-------
		pandas.core.frame.DataFrame
		"""
		dataframe_df = self._obj
		if type(dataframe_df) != pd.core.frame.DataFrame:
			print('The method requires Pandas DataFrame as the input')
			return
		else:
			if col_names_list == None:
				col_names_list = dataframe_df.columns.tolist()
			if not isinstance(col_names_list, list):
				col_names_list = [col_names_list]

			if len(col_names_list) > 0:
				fetcher = []
				for col in col_names_list:
					tupx = self.isnull(col)
					fetcher.append({'Col_Name': col, 'Missing_Count': tupx[0], 'Missing_Percentage': tupx[1]})
				fetcher = pd.DataFrame(fetcher)
				return fetcher
			else:
				print('There is no item in the columns')
				return None


	def nunique(self, col_names_list=None) -> pd.core.frame.DataFrame:
		"""
		To generate:
			1) the number of unqiue value
			2) the percentage of the unique value over the total records (or rows)

		Parameters
		----------
		col_names_list: list
			If not specified, all column names will be used

		Returns
		-------
		pd.core.frame.DataFrame
		"""
		df_input = self._obj
		if col_names_list == None:
			col_names_list = df_input.columns.tolist()
		if len(col_names_list) > 0:
			fetcher = []
			for col in col_names_list:
				n = df_input[col].nunique()
				fetcher.append({'Col_Name': col, 'Unique_Values': n})
			fetcher = pd.DataFrame(fetcher)
			fetcher['Prcent_Unique_Values'] = fetcher.Unique_Values.map(lambda x: x / len(df_input))
			return fetcher


	def ci(self, col, n=1000, func=np.mean, p=0.05):
		"""
		Generate 'n' bootstrap samples, evaluating `func` at each resampling.
		This method returns a function, which can be called to obtain confidence intervals of interest.
		Parameters
		----------
		n: int, optional
			sample size for the sampling distribution
			(defalt = 1,000)

		func: function, optional
			The statistic functions to be bootstrapped its sampling distribution
			(default = np.mean())

		p: float, optional
			p-value for specifyin 2-sided symmetric confidence interval

		Returns
		-------
		function
			Function to be called to obtain confidence intervals of interest.
			Return 2-sided symmetric confidence interval specified
		"""
		df = self._obj.copy()
		data = df[col].copy()
		simulations = list()
		sample_size = len(data)
		xbar_init = np.mean(data)
		for c in range(n):
			itersample = np.random.choice(data, size=sample_size, replace=True)
			simulations.append(func(itersample))
		simulations.sort()

		def ci(p):
			u_pval = (1 + p) / 2.
			l_pval = (1 - u_pval)
			l_indx = int(np.floor(n * l_pval))
			u_indx = int(np.floor(n * u_pval))
			return (simulations[l_indx], simulations[u_indx])
		return (ci)


	def cols_search(self, keywords_list:Union[str, list], method:str='any', case_sensitive=False) -> list:
		"""
		To search the columns of which its name contains the 'any' or 'all' the keywords.
		Single string (str) can also be passed to the keywords_list.

		Parameters
		----------
		keywords_list
		method
		case_sensitive

		Returns
		-------
		list_of_column_names: list
		"""
		df = self._obj

		if not isinstance(keywords_list, Iterable):
			keywords_list = [keywords_list]
		cols = df.columns.tolist().copy()

		if not case_sensitive:
			cols = [c.lower() for c in cols]
			keywords_list = [k.lower() for k in keywords_list]
			print("Non-CaseSensitive")

		fetcher = []
		if method == 'any':
			for col in cols:
				for kw in keywords_list:
					if kw in col:
						fetcher.append(col)
						break
		elif method == 'all':
			for col in cols:
				if all([kw in col for kw in keywords_list]):
					fetcher.append(col)
		return fetcher


	def cols_contain(self, keyword, case_sensitive=False, columns=None, force_string=False):
		"""
		To search for column which its name contain the keyword specified.

		Parameters
		----------
		keyword
		case_sensitive
		columns
		force_string

		Returns
		-------

		"""
		df = self._obj.copy()

		if columns is None:
			columns = df.columns.tolist()

		fetcher = []
		for col in columns:
			try:
				if force_string:
					df[col] = df[col].astype(str)

				if not case_sensitive:
					df[col] = df[col].str.lower()

				status = df[col].str.contains(keyword)
				status = status.any(axis='index')
				fetcher.append([col, status])
			except:
				pass
		fetcher = pd.DataFrame(fetcher, columns=['Column', 'Status'])
		fetcher = fetcher[fetcher.Status == True]
		return list(fetcher.Column)


	def reset_index(self, index_label:str="RID", inplace:bool=True):
		"""
		To reset index and immediately rename the old 'index' to new index_label defined.

		Parameters
		----------
		index_label: str, optional

		inplace: bool, optional

		Returns
		-------
		pd.core.frame.DataFrame
			ONLY when inplace == False
		"""

		df_input = self._obj
		if inplace:
			df = df_input
		else:
			df = df_input.copy()

		df.reset_index(inplace=True)
		df.ds.rename('index', index_label)

		if inplace:
			df_input = df.copy()
		else:
			return df


	def cols_shift(self, col_names:Union[str, list], direction:Union[str, int]= 'right'):
		"""
		To shift a list of columns to the left-most or the right-most of the dataframe.

		Parameters
		----------
		col_names: str or list

		direction: str or int
			str = 'left' or right
			int = 0 or 1

		inplace

		Returns
		-------
		df with reordered columns: pd.core.frame.DataFrame

		"""
		df_input = self._obj.copy()

		if not isinstance(col_names, list):
			col_names = [col_names]

		df_cols_list = df_input.columns.tolist().copy()

		for col in col_names:
			df_cols_list.remove(col)

		if (direction == 'right') | (direction == 1):
			df_cols_list.extend(col_names)
			df_input = df_input[df_cols_list].copy()
		elif (direction == 'left') | (direction == 0):
			col_names.extend(df_cols_list)
			df_input = df_input[col_names].copy()

		return df_input


	def _split(self, col:str, sep:str, index_label:str="RID", drop_innerindex:bool=True):
		'''
		Function to be depreciated
		'''
		warnings.warn('Method to be depreciated', DeprecationWarning, stacklevel=2)
		df = self._obj.copy()
		if index_label is not None:
			df = df[[index_label, col]].copy()
			df.set_index(index_label, inplace=True)

		df = df[col].str.split(sep, expand=True)
		df = pd.DataFrame(df.stack()).reset_index()
		df.columns = [index_label, 'InnerIndex', col]
		df[col] = df[col].str.replace(sep, '')
		df[col] = df[col].str.strip()
		if drop_innerindex:
			df.drop('InnerIndex', 'columns', inplace=True)
		return df


	def split(self, col:str, sep:str, index_label:str='RID',
	                                             drop_innerindex:bool=True, reset_index_inplace:bool=True):
		"""
		To generate a DataFrame by splitting the values in a string, where the values are separated by a separator
		character.

		This method is improved upon the original split method in pandas. Where there is no separator in a row,
		the value will still be posted to the newly generated DataFrame as the outputs.

		Parameters
		----------
		col: str
		sep: str
		index_label: str
		drop_innerindex: bool
		reset_index_inplace

		Returns
		-------
		pd.core.frame.DataFrame
		"""
		if reset_index_inplace:
			dfo = self._obj
		else:
			dfo = self._obj.copy()
		dfo.reset_index(drop=True, inplace=True)
		dfo.ds.reset_index(index_label=index_label)

		df = dfo.copy()
		df[col + "_splitready"] = df[col].map(lambda x: str(x) + sep if sep not in str(x) else str(x))
		df = df[[index_label, col + "_splitready"]].copy()
		df.set_index(index_label, inplace=True)
		df = df[col + "_splitready"].str.split(sep, expand=True).stack().reset_index()
		df.columns = [index_label, 'InnerIndex', col]

		df[col] = df[col].str.replace(sep, '')
		df[col] = df[col].replace('', np.nan)
		df.dropna(subset=[col], inplace=True)

		df[col] = df[col].str.strip()

		if drop_innerindex:
			df.drop('InnerIndex', 'columns', inplace=True)

		return df


	def rename(self, col_index_or_name:Union[str, int], col_name_new, inplace:bool=True):
		"""
		To rename single column
		Parameters
		----------
		col_index_or_name
		col_name_new
		inplace

		Returns
		-------
		renamed_DataFrame: pd.core.frame.DataFrame
			Only if inplace is set to False.
		"""
		df_input = self._obj
		if inplace:
			df = df_input
		else:
			df = df_input.copy()

		if str(col_index_or_name).isnumeric() == True:
			colName_Old = df.columns.tolist()[col_index_or_name]
		else:
			colName_Old = col_index_or_name

		if colName_Old in [col for col in df.columns.tolist()]:
			df.rename(columns={colName_Old: col_name_new}, inplace=True)
		else:
			print("Error: The " + colName_Old + " is not an existing column name.")

		if inplace:
			df_input = df.copy()
		else:
			return df


	def merge(self, right, how='left', on=None, left_on=None, right_on=None, isnull=None) -> pd.core.frame.DataFrame:
		"""
		To merge with another DataFrame.
		A wrapper method for 'merge' in pandas, with additional checking mechanisms.
		The mehtod also creates a backup of the original DataFrame with the key 'last' in dsx.backup_repo (dictionary).

		Parameters
		----------
		right: pd.core.frame.DataFrame
		isnull: str

		Returns
		-------
		pd.core.frame.DataFrame
		"""
		df = self._obj

		self.backup_repo['last'] = df.copy()
		if on is not None:
			df = df.merge(right, how=how, on=on)
		else:
			df = df.merge(right, how=how, left_on=left_on, right_on=right_on)

		print(self.len_compare(df, self.backup_repo['last']))
		if isnull is not None:
			print(df.ds.isnull(isnull))
		return df


	def to_dict(self, key_col:str, val_col:str) -> pd.core.frame.DataFrame:
		"""
		To generate dictionary from two columns
		Parameters
		----------
		key_col: str
		val_col: str

		Returns
		-------
		pd.core.frame.DataFrame
		"""
		df = self._obj
		dict = {}
		for index, row in df.iterrows():
			dict[row[key_col]] = row[val_col]
		return dict


	def _df_convert_date_columns_to_pandas_date(self, keyword='date', cols_list=None, inplace=False):
		"""
		To conver columns with "date" (non-case sensitive) in DataFrame into Pandas DateTimes type.
		If the <<cols_list>> paramater is provided, the <<keyword>> parameter will be ignored

		Parameters
		----------
		keyword
		cols_list
		inplace

		Returns
		-------

		"""
		df_input = self._obj
		if inplace:
			df = df_input
		else:
			df = df_input.copy()

		if cols_list != None:
			if isinstance(cols_list, Iterable):
				date_Cols = cols_list
		else:
			date_Cols = [col for col in df.columns.tolist() if keyword in col.lower()]
		for col in date_Cols:
			try:
				df[col] = pd.to_datetime(df[col])
				print("Converted Column - " + str(col))
			except:
				print("Conversation Failed for Column - " + str(col))

		if not inplace:
			return df


	def len_compare(self, df_to_compare, overwrite_df1=None) -> tuple:
		"""
		Compare the length of two Dataframes (or any other enumeratable object)

		Parameters
		----------
		df_to_compare:

		overwrite_df1: bool, optional
			To ignore this instance of DataFrame and use the DataFrame in parameter as the copy to be compared.

		Returns
		-------

		"""
		if overwrite_df1 is None:
			df = self._obj
		else:
			df = overwrite_df1

		if len(df) == len(df_to_compare):
			return (True, len(df))
		else:
			return (False, len(df), len(df_to_compare))


	def cumsum(self, col_name: str) -> pd.core.frame.DataFrame:
		"""
		 To generates the following using the unique values of a variable:
		 - Count (Raw Count of Records)
		 - Percentage of the values over the total data
		 - Accumulated percentage of the values
		Parameters
		----------
		col_name: str

		Returns
		-------
		pd.core.frame.DataFrame
		"""
		df = self._obj.copy()
		#df = df.fillna('Missing_Value')
		innerTemp = pd.DataFrame(df[col_name].value_counts()).reset_index()
		innerTemp.sort_values(['index'], 'index', inplace=True)
		innerTemp['Records_Percent'] = innerTemp[col_name] / len(df)
		innerTemp.sort_values('Records_Percent', 0, False, inplace=True)
		innerTemp['Accum_Percent'] = innerTemp['Records_Percent'].cumsum()
		innerTemp.columns = [col_name, 'Records_Count', 'Records_Percent', 'Accum_Percent']
		return innerTemp


	def to_excel_exists(self, filepath_incl_extension, tab_name, index=False):
		"""
		To insert a DataFrame into a new worksheet in an existing excel file.
		The method use 'openpyxl' as the writer engine.

		Parameters
		----------
		filepath_incl_extension: str
		tab_name: str
		index: bool

		Returns
		-------

		"""
		from openpyxl import load_workbook
		book = load_workbook(filepath_incl_extension)
		writer = pd.ExcelWriter(filepath_incl_extension, engine='openpyxl', index=index)
		writer.book = book
		self._obj.to_excel(writer, sheet_name=tab_name)
		writer.save()
		writer.close()


	def to_excel_stringify(self, dir=None, strings_to_urls_bool=False):
		"""
		Faster option to export Excel File, with the option to stringify all hyperlinks in the table.
		Parameters
		----------
		dir
		strings_to_urls_bool

		Returns
		-------

		"""
		if dir is None:
			dir = self.dir_data

		writer = pd.ExcelWriter(dir, engine='xlsxwriter', options={'strings_to_urls': strings_to_urls_bool})
		self._obj.to_excel(writer)
		writer.close()
		print("Exported Excel File to " + dir)


	def to_excel(self, filename: str, dir=None, extension='xlsx', index=True, index_label="RID", freeze_panes=(1,0)):
		"""
		Quick export to Excel. No file extension required.
		Parameters
		----------
		filename
		dir
		extension
		index
		index_label
		freeze_panes

		Returns
		-------
		None
		"""

		if dir is None:
			dir = self.dir_data

		if index:
			self._obj.to_excel(os.path.join(dir, filename + "." + extension), index=True, index_label=index_label, freeze_panes=freeze_panes)
		else:
			self._obj.to_excel(os.path.join(dir, filename + "." + extension), freeze_panes=freeze_panes)
		print("Exported Excel File to " + dir)


	def dump(self, path:str, compression_level:int=7):
		"""
		To dump DataFrame to the project's data/temp directory

		Parameters
		----------
		path: str
		dir: str, optional
			Default = data/temp

		compression_level: int, optional

		Returns
		-------
		None
		"""
		df = self._obj

		if '/' not in path:
			dir = os.path.join(self.dir_data, 'temp', path)
		else:
			dir = path

		dump(df, dir + ".df", compress=('gzip', compression_level))
		print("Compressed dump created at " + os.path.dirname(dir) + " | filename = " + os.path.basename(dir + ".df"))


	def to_excel_fast(self, filename, sheet_name='Sheet1', origin=(1, 1), columns=True, index=False):
		"""
		Write DataFrame to excel file using pyexelerate library. Fast writing, with data integrity traded-off
		"""
		import pyexcelerate

		df = self._obj

		if not isinstance(filename, pyexcelerate.Workbook):
			location = filename
			filename = pyexcelerate.Workbook()
		else:
			location = None
		worksheet = filename.new_sheet(sheet_name)

		# Account for space needed for index and column headers
		column_offset = 0
		row_offset = 0

		if index:
			index = df.index.tolist()
			ro = origin[0] + row_offset
			co = origin[1] + column_offset
			worksheet.range((ro, co), (ro + 1, co)).value = [['Index']]
			worksheet.range((ro + 1, co), (ro + 1 + len(index), co)).value = list(map(lambda x: [x], index))
			column_offset += 1
		if columns:
			columns = df.columns.tolist()
			ro = origin[0] + row_offset
			co = origin[1] + column_offset
			worksheet.range((ro, co), (ro, co + len(columns))).value = [[*columns]]
			row_offset += 1

		# Write the data
		row_num = df.shape[0]
		col_num = df.shape[1]
		ro = origin[0] + row_offset
		co = origin[1] + column_offset
		worksheet.range((ro, co), (ro + row_num, co + col_num)).value = df.values.tolist()

		if location:
			filename.save(location)
			print("Successfully exported to " + str(location))


	def cols_datetime_to_string(self, inplace=False):
		df_input = self._obj.copy()
		if inplace:
			df = df_input
		else:
			df = df_input.copy()
		dtypex = pd.DataFrame(df.dtypes).reset_index()
		dtypex.columns = ['Colname', 'Type']
		dtypex.Type = dtypex.Type.astype(str)
		dtypex = dtypex[dtypex.Type.str.contains('date')]
		for col in dtypex.Colname.tolist():
			df[col] = df[col].astype(str)
			print("Converted {}".format(col))

		if inplace:
			df_input = df.copy()
		else:
			return df


	def convert_dtypes(self):
		"""
		To convert dtypes to Pandas 1.0 dtypes and stringify object columns

		Returns
		-------
		pd.core.frame.DataFrame
		"""
		df = self._obj.copy()
		df = df.convert_dtypes()
		df_types = pd.DataFrame(df.dtypes).reset_index()
		df_types.columns = ['colname', 'datatype']
		df_types.datatype = df_types.datatype.astype(str)
		df_types = df_types[df_types.datatype.str.contains('datetime')]
		if len(df_types):
			for col in df_types.colname:
				df[col] = df[col].astype(str)
		return df.copy()


	def hash(self, key:str, salt:str=None, reverse=False, parallel:bool=False, parallel_statusbar:bool=False, inplace=False):
		"""
		To hash an entire DataFrame
		Requires itsdangerous package. Install using:
			"pip install itsdangerous"
		Parameters
		----------
		key: str
		salt: str, optional
		parallel: bool, optional
		parallel_statusbar: bool, optional

		Returns
		-------
		pd.core.frame.DataFrame
		"""
		df_input = self._obj.copy()
		if inplace:
			df = df_input
		else:
			df = df_input.copy()

		from itsdangerous.url_safe import URLSafeSerializer
		serializer = URLSafeSerializer(key, salt=salt)
		if reverse:
			serializer_func = serializer.loads
		else:
			serializer_func = serializer.dumps

		if parallel:
			from pandarallel import pandarallel
			pandarallel.initialize(progress_bar=parallel_statusbar)
			df = df.parallel_applymap(lambda x: serializer_func(x))
		else:
			df = df.applymap(serializer_func)

		df.columns = [serializer_func(c) for c in df.columns.tolist()]

		if inplace:
			df_input = df.copy()
		else:
			return df


	def to_xv(self, title=None, convert_time:bool=True, dirbase=None):
		#import webbrowser
		df = self._obj
		if convert_time:
			df = df.ds.cols_datetime_to_string(inplace=False)

		if title is None:
			title = 'LabView'
			html_filename = 'LabView'
		else:
			html_filename = 'LabView_' + str(title)

		# Writing data to disk
		jsonstr = df.to_json(orient='records')
		data_var = "var data = {};".format(jsonstr)


		base_file = os.path.join(dirbase, 'xbase.html')
		htmlfile = open(base_file, 'r')
		htmlstring = htmlfile.read()
		htmlfile.close()

		# In the <script> tag which points to the external data.js
		#htmlstring = htmlstring.replace('data.js', str(data_filename) + '.js', 1)
		htmlstring = htmlstring.replace('LabView', str(title), 1)
		htmlstring = htmlstring.replace('e="insert_data_here";', data_var, 1)


		newfile = open(os.path.join(dirbase, html_filename + '.html'), 'w')
		newfile.write(htmlstring)
		newfile.close()
		return html_filename
		#webbrowser.get(dsx.path_chrome).open(path_string)


	def xv(self, title=None, convert_time=True, width="100%", height="1200", dirhtml="../_temp", dirbase="_temp", **kwargs):
		"""

		Parameters
		----------
		title: str, Title for the new viewer file.

		convert: bool, Convert datetime dtype to str for display.


		Returns
		-------

		"""
		from IPython.display import IFrame
		viewer_filename = self._obj.ds.to_xv(title, convert_time=convert_time, dirbase=dirbase)
		return IFrame(os.path.join(dirbase, (viewer_filename+'.html')), width=width, height=height)



	@classmethod
	def _modify_vizdatafile(cls, data_filename='data', viewer_filename=None, dir_file=None):
		"""
		Depecirated Method
		A private classmethod for changing the data.js configuration in the js and the html files.

		Parameters
		----------
		filename
		dir_file

		Returns
		-------

		"""
		if dir_file is None:
			dir_file = '_temp/xbox.html'
		htmlfile = open(dir_file, 'r')
		htmlstring = htmlfile.read()
		htmlfile.close()

		# In the <script> tag which points to the external data.js
		htmlstring = htmlstring.replace('data.js', str(data_filename) + '.js', 1)
		htmlstring = htmlstring.replace('title_to_replace', str(data_filename), 1)

		dir_name = os.path.dirname(dir_file)
		newfile = open(os.path.join(dir_name, data_filename + '.html'), 'w')
		newfile.write(htmlstring)
		newfile.close()



	def bk(self, bk_name: str = None):
		"""
		To backup the dataframe.

		Parameters
		----------
		bk_name

		Returns
		-------
		None
		"""
		if bk_name is not None:
			self.backup_repo[str(bk_name)] = self._obj.copy()
		else:
			try:
				name = self._obj.ds.get_dfname()
				self.backup_repo[str(name)] = self._obj.copy()
			except:
				next_index = len(self.backup_repo)
				self.backup_repo["df_" + str(next_index)] = self._obj.copy()
				self._obj.name = "df_" + str(next_index)
				print("Backup to {:s}".format("df_" + str(next_index)))


	def rs(self, bk_name: str = None, inplace=True):
		"""
		To restore the dataframe.

		Parameters
		----------
		bk_name
		inplace

		Returns
		-------
		pandas.core.frame.DataFrame
		"""
		df = self._obj
		if bk_name is not None:
			df = self.backup_repo[str(bk_name)].copy()
		else:
			df = self.backup_repo[self._obj.name].copy()
			print("Restored from {:s}".format(self._obj.name))
		if not inplace:
			return df


	@classmethod
	def backup(cls, df, name:str='last'):
		"""
		To backup the DataFrame or List (or any object with .copy() method)

		Parameters
		----------
		df
			DataFrame or List (or any object with .copy() method)

		name
			Name of the backup. To be used to retrieve the data.

		Returns
		-------
		None
		"""

		cls.backup_repo [name] = df.copy()



	@classmethod
	def restore(cls, name:str='last'):
		"""
		To restore the DataFrame or List (or any object with .copy() method)

		Parameters
		----------
		df
			DataFrame or List (or any object with .copy() method)

		name
			Name of the backup. To be used to retrieve the data.

		Returns
		-------
		Object
		"""
		if name is not None:
			return cls.backup_repo[name].copy()
		else:
			return cls.backup_repo['last'].copy()



	@staticmethod
	def delta_todate(num_yyyyddd):
		"""
		To convert timedelta to date
		Parameters
		----------
		num_yyyyddd

		Returns
		-------
		datetime.datetime
		"""
		import datetime
		str_yyddd = str(num_yyyyddd)
		year = int(str_yyddd[:4])
		baseDate = datetime.date(year, 1, 1)
		days = int(str_yyddd[4:])
		return baseDate + datetime.timedelta(days=days)


	@staticmethod
	def to_numeric(inputString):
		"""
		To convert string to numeric
		Parameters
		----------
		inputString

		Returns
		-------

		"""
		numList = [x for x in inputString if x.isdigit()]
		return ''.join(numList)


	@staticmethod
	def progress(iterable:Iterable, counter:int) -> str:
		"""
		To return string template for the progress of a loop operation.

		Parameters
		----------
		iterable: Iterable
		counter: int

		Returns
		-------
		str
		"""
		progress = counter / len(iterable) * 100
		return str("Processing record " + str(counter) + " over " + str(len(iterable)) + " | " + str(
			np.round(progress, 4)) + '%')


	def get_dfname(self, set=True):
		"""
		To get name of the variable.


		Only work in iPython.

		Parameters
		----------
		var

		Returns
		-------
		variable_name: str
		"""
		callers_local_vars = inspect.currentframe().f_back.f_locals.items()
		name = [var_name for var_name, var_val in callers_local_vars if var_val is self._obj][0]
		if set:
			self._obj.name = name
		return name


	@staticmethod
	def get_varname(var:object):
		"""
		To get name of the variable.

		Only work in iPython.

		Parameters
		----------
		var

		Returns
		-------
		variable_name: str
		"""
		callers_local_vars = inspect.currentframe().f_back.f_locals.items()
		return [var_name for var_name, var_val in callers_local_vars if var_val is var][0]


	@staticmethod
	def interactive():
		"""
		Set InteractiveShell.ast_node_interactivity = "all"
		Set mpl.use("module://backend_interagg")
		Set plt.ion()

		"""
		from IPython.core.interactiveshell import InteractiveShell
		InteractiveShell.ast_node_interactivity = "all"
		mpl.use("module://backend_interagg")
		plt.ion()
		print("Package loaded in Non-Notebook Mode | mpl.use('module://backend_interagg') | plt.ion()" )


	@staticmethod
	def plt_labels(percent=False, fontsize=None, color=None, denominator=None):
		"""
		To insert label for each element in the current axes (last chart created).
		Parameters
		----------
		percent: bool
		fontsize: float
		color: str
		denominator: float

		Returns
		-------
		None
		"""
		ax = plt.gca()
		for p in ax.patches:
			if percent:
				ax.text(p.get_x() + (p.get_width() / 2), p.get_height(), round(p.get_height() / denominator * 100, 1),
				        ha='center', va='bottom')
			else:
				ax.text(p.get_x() + (p.get_width() / 2), p.get_height(), p.get_height(), ha='center', va='bottom')



	@classmethod
	def activate_lolviz(cls):
		"""
		Import lolviz package as lz.
		Add graphviz directory to the os.environ["path"].

		Parameters
		----------
		lolviz_dir: str, optional

		Returns
		-------
		lolviz instance
		"""

		# lolviz_dir='C:/Program Files (x86)/Graphviz2.38/bin/'
		import lolviz as lz
		os.environ["PATH"] += os.pathsep + dsx.path_graphviz
		print('Activate lolviz with path {}'.format(dsx.path_graphviz))
		return lz


	@classmethod
	def set_path_chrome(cls, path_string:str):
		"""
		Set the path to web browser's (prefer: Chrome) executable
		Returns
		-------

		"""
		cls.path_chrome = path_string


	@classmethod
	def set_path_graphviz(cls, path_string:str):
		"""
		Set the path to Graphviz's executable
		Returns
		-------

		"""
		cls.path_graphviz = path_string


	@classmethod
	def set_dirs(cls, root=False):
		"""
		Set the project root folder.

		Parameters
		----------
		root: bool, optional
			To indicate whether the current active directory is the root or sub-directory of the project

		Returns
		-------
		None
		"""

		dir = None
		if root:
			dir = os.getcwd()
		else:
			dir = os.path.join(os.getcwd(), "..")
		os.chdir(dir)

		cls.dir_project = os.getcwd()
		cls.dir_data = os.path.join(cls.dir_project, 'data')
		cls.dir_temp = os.path.join(cls.dir_project, '_temp')

		print('Set project directory to {}.'.format(os.getcwd()))
		print('Property "dir_%" enabled')


	@classmethod
	def setup_project(cls, root=True, get_xfiles=False, git_files=False):
		"""
		Setup project directories for new projects.
		If the directories exist, will not be overwritten.

		Parameters
		----------
		root: bool, optional
		get_xfiles: bool, optional
		git_files: bool, optional

		Returns
		-------
		None
		"""

		folders = ['data', 'data/temp', 'data/inputs', 'data/outputs', 'notebooks', '_temp']
		for folder in folders:
			if os.path.exists(os.path.join(cls.dir_project, folder)) == False:
				os.mkdir(os.path.join(cls.dir_project, folder))
		print('Created project structure')

		cls.set_dirs(root=root)

		if get_xfiles:
			os.mkdir(os.path.join(cls.dir_project, '_temp'))

			os.chdir(cls.dir_temp)
			if not os.path.exists('xbase.html'):
				import urllib
				urllib.request.urlretrieve('https://s3.ap-southeast-1.amazonaws.com/nictsyen/shared/xbase.html', 'xbase.html')


			os.chdir(cls.dir_project)
			print('Downloaded Extra Files.')


	@classmethod
	def del_tempfiles(cls, tempdata=False):
		"""
		Static method: To delete temporary files of the projects.

		Parameters
		----------
		tempdata: bool, optional
			Default is 'False'. Set to 'True' to delete temporary data in 'data/temp' directory

		Returns
		-------
		None
		"""
		import shutil
		try:
			path = os.path.join(cls.dir_data, 'temp')
			if os.path.exists(path):
				shutil.rmtree(path)
				print('Deleted data/temp')

			if os.path.exists(cls.dir_temp):
				shutil.rmtree(cls.dir_temp)
				print('Deleted _temp')
		except:
			print("Removed files. Directories may remain")


	@staticmethod
	def matplotlib_config():
		"""
		Print matplotlib configurations

		Returns
		-------
		lines of texts: str

		"""
		print('%matplotlib inline')
		print("%config InlineBackend.figure_format = 'retina'")
		#print("sns.set_style('fivethirtyeight')")
		print("plt.rc('figure', figsize=(16,9))")
		print("sns.set_context(context={'figure.figsize': (16,9)})")
		print("plt.style.use('fivethirtyeight')")

	@staticmethod
	def qgrid_config():
		print("qgrid.set_grid_option('forceFitColumns', False)")


	@staticmethod
	def set_ipython(node_interactivity:str='last'):
		"""
		Set ast_node_interactivity in Ipython.core.InteractiveShell

		Parameters
		----------
		node_interactivity: str, optional
			Default is 'last'. DSX uses 'all' if kernel is detected.

		Returns
		---
		None

		"""
		from IPython.core.interactiveshell import InteractiveShell
		InteractiveShell.ast_node_interactivity = node_interactivity


if __name__ != '__main__':
	print('Successfully imported ds_utils as Package')
else:
	print('Successfully executed ds_utils. Warning! - This is not pd_utils is intended to be run as a script.')